import requests
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import time
import datetime
from datetime import date


def trackingIdDecryption(trackingNumber):
    trackingIdEncryption = {'A': '%01', 'B': '%02', 'C': '%03', 'D': '%04', 'E': '%05', 'F': '%06', 'G': '%07',
                            'H': '%08', 'I': '%09', 'J': '%0A', 'K': '%0B', 'L': '%0C', 'M': '%0D', 'N': '%0E',
                            'O': '%0F', 'P': '%10', 'Q': '%11', 'R': '%12', 'S': '%13', 'T': '%14', 'U': '%15',
                            'V': '%16', 'W': '%17', 'X': '%18', 'Y': '%19', 'Z': '%1A',
                            '0': 'n', '1': 'o', '2': 'p', '3': 'q', '4': 'r', '5': 's', '6': 't', '7': 'u',
                            '8': 'v', '9': 'w'}
    return "".join([trackingIdEncryption[letter] for letter in trackingNumber])

def readExcelFile(startRow, endRow):
    # get handle on existing file
    wb = load_workbook(filename=os.environ['EXCEL_SHEET'])
    # get current month and year worksheet
    month_year = date.today().strftime("%b-%Y").split("-")
    worksheet = "{month}. {year}".format(month=month_year[0], year=month_year[1])
    # print(worksheet)
    ws = wb[worksheet]
    font = Font(name='Times New Roman', size=12)
    alignment = Alignment(horizontal='center')
    # get columns
    # December 2020
    # shipping_status_col = 'O'
    # time_col = 'P'
    # tracking_number_col = 'M'
    # January 2021 and on
    shipping_status_col = 'N'
    time_col = 'O'
    tracking_number_col = 'P'
    # loop through range values
    for i in range(startRow, endRow+1):
        cell = tracking_number_col + str(i)
        if isinstance(ws[cell].value, str):
            # track only orders that have not been delivered yet
            if ws[shipping_status_col + str(i)].value != 'Delivered':
                print('Row ' + str(i) + ' Tracking Number: ' + ws[cell].value)
                status = getShippingStatus(ws[cell].value)
                if status is not None: ws[shipping_status_col + str(i)] = status
                ws[time_col + str(i)] = datetime.datetime.now()
            else: continue
        else:
            ws[shipping_status_col + str(i)] = 'No Tracking Number'
            ws[shipping_status_col + str(i)] = 'No Tracking Number'
        ws[shipping_status_col + str(i)].font = font
        ws[shipping_status_col + str(i)].alignment = alignment
        ws[time_col + str(i)].font = font
        ws[time_col + str(i)].alignment = alignment
        print(ws[shipping_status_col + str(i)].value)
        print(ws[time_col + str(i)].value)

    wb.save(filename=os.environ['EXCEL_SHEET'])


def getShippingStatus(trackingNum):
    request_url = os.environ['TRACKING_API']
    url = f"{os.environ['TRACKING_SITE']}{trackingNum}"
    trackingId = trackingIdDecryption(trackingNum)
    print(trackingId)
    data = {
        'trackingId': trackingId,
        'carrier': 'Auto-Detect',
        'language': 'en',
        'country': 'Russian Federation',
        'platform': 'web-desktop',
        'wd': 'false',
        'c': 'false',
        'p': 3,
        'l': 2,
        'se': '1792x1017,MacIntel,Gecko,Mozilla,Netscape,Google Inc.,4g,Intel Inc.,Intel(R) UHD Graphics 630,undefined,103,3584,2034'}
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36'}
    try:
        resp = requests.post(request_url, data=data, headers=headers)
        print(resp)
        print(resp.json())
        if 'error' in resp.json():
            print('Error: ' + resp.json()['error'].title())
            return 'Error: ' + resp.json()['error'].title()
        elif 'sub_status' in resp.json():
            print('Sub-status: ' + resp.json()['sub_status'].title())
            return resp.json()['sub_status'].title()
        else:
            print('Status: ' + resp.json()['status'].title())
            return resp.json()['status'].title()
    # check to see if request does not give back a valid json
    except ValueError:
        print('Too many requests, the request did not return any JSON')
        return None
    # except ConnectionError:
    #     print(ConnectionError)
    #     return None


if __name__ == '__main__':
    try:
        # startRow = int(input('Please input the first row you would like to track: '))
        # endRow = int(input('Please input the last row you would like to track up to: '))
        startRow = 2
        endRow = 4
        # last row is 2247
        start = time.time()
        readExcelFile(startRow, endRow)
        end = time.time()
        # total time taken
        print(f"Runtime of the program is {end - start}")
    except ValueError as ve:
        print('Error Message: ' + str(ve))
        print('Please input valid integers for the start and end rows.')
