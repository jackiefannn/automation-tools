import asyncio
import datetime
import sys
import time
from datetime import date
import aiohttp
from aiohttp import ClientSession
import logging
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# initialize basic logging information for us to use
logging.basicConfig(
    format="%(asctime)s %(levelname)s:%(name)s: %(message)s",
    level=logging.DEBUG,
    datefmt="%H:%M:%S",
    stream=sys.stderr,
)
logger = logging.getLogger("areq")
logging.getLogger("chardet.charsetprober").disabled = True


def trackingIdDecryption(trackingNumber):
    trackingIdEncryption = {'A': '%01', 'B': '%02', 'C': '%03', 'D': '%04', 'E': '%05', 'F': '%06', 'G': '%07',
                            'H': '%08', 'I': '%09', 'J': '%0A', 'K': '%0B', 'L': '%0C', 'M': '%0D', 'N': '%0E',
                            'O': '%0F', 'P': '%10', 'Q': '%11', 'R': '%12', 'S': '%13', 'T': '%14', 'U': '%15',
                            'V': '%16', 'W': '%17', 'X': '%18', 'Y': '%19', 'Z': '%1A',
                            '0': 'n', '1': 'o', '2': 'p', '3': 'q', '4': 'r', '5': 's', '6': 't', '7': 'u',
                            '8': 'v', '9': 'w'}
    return "".join([trackingIdEncryption[letter] for letter in trackingNumber])


async def fetch_html(url: str, session: ClientSession, data, **kwargs) -> str:
    """
    GET request wrapper to fetch page HTML.
    kwargs are passed to `session.request()`.
    """
    headers = {'content_type': 'text/html', 'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X x.y; rv:42.0) Gecko/20100101 Firefox/42.0'}
    try:
        resp = await session.request(method="POST", url=url, data=data, headers=headers, **kwargs)
        resp.raise_for_status()
        logger.info("Got response [%s] for URL: %s", resp.status, url)
        doc = await resp.json(content_type=None)
        logger.info("JSON Response of request: %s", doc)
    except Exception as ve:
        logger.error("Too many requests, the request did not return any JSON and returned nothing: %s", getattr(ve, "__dict__", {}))
    else:
        return doc


async def fetch_page(url: str, session: ClientSession, data, **kwargs):
    """Ensure that request was able to fetch JSON response for `url`."""
    try:
        doc = await fetch_html(url=url, session=session, data=data, **kwargs)
    except (aiohttp.ClientError, aiohttp.http_exceptions.HttpProcessingError) as e:
        logger.error("aiohttp exception for %s [%s]: %s", url, getattr(e, "status", None), getattr(e, "message", None))
        return None
    except Exception as e:
        logger.exception("Non-aiohttp exception occured:  %s", getattr(e, "__dict__", {}))
        return None
    else:
        return doc


async def fetch_and_write_one(url: str, database, session, data, row, **kwargs):
    """ Fetch JSON doc and return status of shipping order """
    doc = await fetch_page(url=url, session=session, data=data, **kwargs)
    if not doc: return None
    # logger.info(f"Contents of the JSON response: {doc}")
    status = str
    if 'error' in doc:
        logger.info('Error: ' + doc['error'].title())
        status = f"Error: {doc['error'].title()}"
    elif 'sub_status' in doc:
        logger.info('Sub-status: ' + doc['sub_status'].title())
        status = doc['sub_status'].title()
    else:
        logger.info('Status: ' + doc['status'].title())
        status = doc['status'].title()

    db_res = await as_insert_one(database=database, status=status, row=row)

    return db_res


async def fetch_and_write(url: str, database, **kwargs) -> None:
    """Fetch & write concurrently to `file` for multiple post requests to 'url'."""
    global row
    async with ClientSession() as session:
        post_tasks = []
        while ws[f"{sale_order_col}{row}"].value:
            # track sales orders that have a shipping number
            if isinstance(ws[f"{tracking_number_col}{row}"].value, str):
                # track only orders that have not been delivered yet
                if ws[f"{shipping_status_col}{row}"].value != 'Delivered' and \
                        ws[f"{shipping_status_col}{row}"].value != 'Arrived' and \
                        ws[f"{shipping_status_col}{row}"].value != 'Pickup':
                    trackingNum = ws[f"{tracking_number_col}{row}"].value
                    logger.info(f"Row: {row} & Tracking num: {trackingNum}")
                    logger.info(f"Current Status: {ws[shipping_status_col + str(row)].value}")
                    trackingId = trackingIdDecryption(trackingNum)
                    data = {
                        'trackingId': trackingId,
                        'carrier': 'Auto-Detect',
                        'language': 'en',
                        'country': 'Russian Federation',
                        'platform': 'web-desktop',
                        'wd': 'false',
                        'c': 'false',
                        'p': 3,
                        'l': 2,
                        'se': '1792x1024,MacIntel,Gecko,Mozilla,Netscape,Google Inc.,4g,Intel Inc.,Intel(R) UHD Graphics 630,undefined,103,3584,2048'}

                    res = fetch_and_write_one(url=url, database=database, session=session, data=data, row=row, **kwargs)
                    # if not res: continue
                    post_tasks.append(res)

                    # execute x number of async requests at a time (no need to execute them all at the same time)
                    if len(post_tasks) > 50:
                        await asyncio.gather(*post_tasks)
                        post_tasks = []
            row += 1
        # now execute the rest of the remaining requests
        await asyncio.gather(*post_tasks)


async def as_insert_one(database, status, row):
    ws[f"{shipping_status_col}{row}"] = status
    ws[f"{time_col}{row}"] = datetime.datetime.now()
    ws[f"{shipping_status_col}{row}"] .font = font
    ws[f"{shipping_status_col}{row}"] .alignment = alignment
    ws[f"{time_col}{row}"] .font = font
    ws[f"{time_col}{row}"] .alignment = alignment
    try:
        result = wb.save(filename=database)
        # result = await wb.save(filename=database)
    except TypeError as e:
        logger.error("TypeError exception occured while attempting to save Excel workbook: %s", e)
    logger.info(f"Wrote {status} into row {row}")



if __name__ == '__main__':
    start = time.time()
    # total time taken
    url = os.environ['TRACKING_API']
    database = os.environ['EXCEL_SHEET']
    # get handle on existing file
    wb = load_workbook(filename=os.environ['EXCEL_SHEET'])
    # get current month and year worksheet
    month_year = date.today().strftime("%b-%Y").split("-")
    worksheet = "{month}. {year}".format(month=month_year[0], year=month_year[1])
    ws = wb[worksheet]
    font = Font(name='Times New Roman', size=12)
    alignment = Alignment(horizontal='center')
    row = 2
    sale_order_col = 'A'
    shipping_status_col = 'N'
    time_col = 'O'
    tracking_number_col = 'P'
    asyncio.run(fetch_and_write(url=url, database=database))
    end = time.time()
    logger.info(f"Runtime of the program is {end - start}")
    # asyncio.get_event_loop().run_until_complete(fetch_and_write(url=url, database=database))
