import os
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, numbers


country_abbreviations = {'AF': 'Afghanistan', 'AL': 'Albania', 'DZ': 'Algeria', 'AX': 'Aland Islands',
    'AS': 'American Samoa', 'AI': 'Anguilla', 'AD': 'Andorra', 'AO': 'Angola', 'AN': 'Antilles - Netherlands',
    'AG': 'Antigua and Barbuda', 'AQ': 'Antarctica', 'AR': 'Argentina', 'AM': 'Armenia', 'AU': 'Australia',
    'AT': 'Austria', 'AW': 'Aruba', 'AZ': 'Azerbaijan', 'BA': 'Bosnia and Herzegovina', 'BB': 'Barbados',
    'BD': 'Bangladesh', 'BE': 'Belgium', 'BF': 'Burkina Faso', 'BG': 'Bulgaria', 'BH': 'Bahrain', 'BI': 'Burundi',
    'BJ': 'Benin', 'BM': 'Bermuda', 'BN': 'Brunei Darussalam', 'BO': 'Bolivia', 'BR': 'Brazil', 'BS': 'Bahamas',
    'BT': 'Bhutan', 'BV': 'Bouvet Island', 'BW': 'Botswana',  'BV': 'Belarus', 'BZ': 'Belize', 'KH': 'Cambodia',
    'CM': 'Cameroon', 'CA': 'Canada', 'CV': 'Cape Verde', 'CF': 'Central African Republic', 'TD': 'Chad',
    'CL': 'Chile', 'CN': 'China', 'CX': 'Christmas Island', 'CC': 'Cocos (Keeling) Islands', 'CO': 'Colombia',
    'CG': 'Congo', 'CI': 'Cote D\'Ivoire (Ivory Coast)', 'CK': 'Cook Islands', 'CR': 'Costa Rica',
    'HR': 'Croatia (Hrvatska)', 'CU': 'Cuba', 'CY': 'Cyprus', 'CZ': 'Czech Republic',
    'CD': 'Democratic Republic of the Congo', 'DJ': 'Djibouti', 'DK': 'Denmark', 'DM': 'Dominica',
    'DO': 'Dominican Republic', 'EC': 'Ecuador', 'EG': 'Egypt', 'SV': 'El Salvador', 'TP': 'East Timor',
    'EE': 'Estonia', 'GQ': 'Equatorial Guinea', 'ER': 'Eritrea', 'ET': 'Ethiopia', 'FI': 'Finland', 'FJ': 'Fiji',
    'FK': 'Falkland Islands (Malvinas)', 'FM': 'Federated States of Micronesia', 'FO': 'Faroe Islands',
    'FR': 'France', 'FX': 'France, Metropolitan', 'GF': 'French Guiana', 'PF': 'French Polynesia', 'GA': 'Gabon',
    'GM': 'Gambia', 'DE': 'Germany', 'GH': 'Ghana', 'GI': 'Gibraltar', 'GB': 'Great Britain (UK)', 'GD': 'Grenada',
    'GE': 'Georgia', 'GR': 'Greece', 'GL': 'Greenland', 'GN': 'Guinea', 'GP': 'Guadeloupe',
    'GS': 'S. Georgia and S. Sandwich Islands', 'GT': 'Guatemala', 'GU': 'Guam', 'GW': 'Guinea-Bissau',
    'GY': 'Guyana', 'HK': 'Hong Kong', 'HM': 'Heard Island and McDonald Islands', 'HN': 'Honduras', 'HT': 'Haiti',
    'HU': 'Hungary', 'ID': 'Indonesia', 'IE': 'Ireland', 'IL': 'Israel', 'IN': 'India',
    'IO': 'British Indian Ocean Territory', 'IQ': 'Iraq', 'IR': 'Iran', 'IT': 'Italy', 'JM': 'Jamaica',
    'JO': 'Jordan', 'JP': 'Japan', 'KE': 'Kenya', 'KG': 'Kyrgyzstan', 'KI': 'Kiribati', 'KM': 'Comoros',
    'KN': 'Saint Kitts and Nevis', 'KP': 'Korea (North)', 'KR': 'Korea (South)', 'KW': 'Kuwait',
    'KY': 'Cayman Islands', 'KZ': 'Kazakhstan', 'LA': 'Laos', 'LB': 'Lebanon', 'LC': 'Saint Lucia',
    'LI': 'Liechtenstein', 'LK': 'Sri Lanka', 'LR': 'Liberia', 'LS': 'Lesotho', 'LT': 'Lithuania',
    'LU': 'Luxembourg', 'LV': 'Latvia', 'LY': 'Libya', 'MK': 'Macedonia', 'MO': 'Macao', 'MG': 'Madagascar',
    'MY': 'Malaysia', 'ML': 'Mali', 'MW': 'Malawi', 'MR': 'Mauritania', 'MH': 'Marshall Islands', 'MQ': 'Martinique',
    'MU': 'Mauritius', 'YT': 'Mayotte', 'MT': 'Malta', 'MX': 'Mexico', 'MA': 'Morocco', 'MC': 'Monaco',
    'MD': 'Moldova', 'MN': 'Mongolia', 'MM': 'Myanmar', 'MP': 'Northern Mariana Islands', 'MS': 'Montserrat',
    'MV': 'Maldives', 'MZ': 'Mozambique', 'NA': 'Namibia', 'NC': 'New Caledonia', 'NE': 'Niger',
    'NF': 'Norfolk Island', 'NG': 'Nigeria', 'NI': 'Nicaragua', 'NL': 'Netherlands', 'NO': 'Norway', 'NP': 'Nepal',
    'NR': 'Nauru', 'NU': 'Niue', 'NZ': 'New Zealand (Aotearoa)', 'OM': 'Oman', 'PA': 'Panama', 'PE': 'Peru',
    'PG': 'Papua New Guinea', 'PH': 'Philippines', 'PK': 'Pakistan', 'PL': 'Poland',
    'PM': 'Saint Pierre and Miquelon', 'CS': 'Serbia and Montenegro', 'PN': 'Pitcairn', 'PR': 'Puerto Rico',
    'PS': 'Palestinian Territory', 'PT': 'Portugal', 'PW': 'Palau', 'PY': 'Paraguay', 'QA': 'Qatar', 'RE': 'Reunion',
    'RO': 'Romania', 'RU': 'Russian Federation', 'RW': 'Rwanda', 'SA': 'Saudi Arabia', 'WS': 'Samoa',
    'SH': 'Saint Helena', 'VC': 'Saint Vincent and the Grenadines', 'SM': 'San Marino', 'ST': 'Sao Tome and Principe',
    'SN': 'Senegal', 'SC': 'Seychelles', 'SL': 'Sierra Leone', 'SG': 'Singapore', 'SK': 'Slovakia', 'SI': 'Slovenia',
    'SB': 'Solomon Islands', 'SO': 'Somalia', 'ZA': 'South Africa', 'ES': 'Spain', 'SD': 'Sudan', 'SR': 'Suriname',
    'SJ': 'Svalbard and Jan Mayen', 'SE': 'Sweden', 'CH': 'Switzerland', 'SY': 'Syria', 'SU': 'USSR (former)',
    'SZ': 'Swaziland', 'TW': 'Taiwan', 'TZ': 'Tanzania', 'TJ': 'Tajikistan', 'TH': 'Thailand', 'TL': 'Timor-Leste',
    'TG': 'Togo', 'TK': 'Tokelau', 'TO': 'Tonga', 'TT': 'Trinidad and Tobago', 'TN': 'Tunisia', 'TR': 'Turkey',
    'TM': 'Turkmenistan', 'TC': 'Turks and Caicos Islands', 'TV': 'Tuvalu', 'UA': 'Ukraine', 'UG': 'Uganda',
    'AE': 'United Arab Emirates', 'UK': 'United Kingdom', 'US': 'United States',
    'UM': 'United States Minor Outlying Islands', 'UY': 'Uruguay', 'UZ': 'Uzbekistan', 'VU': 'Vanuatu',
    'VA': 'Vatican City State', 'VE': 'Venezuela', 'VG': 'Virgin Islands (British)', 'VI': 'Virgin Islands (U.S.)',
    'VN': 'Vietnam', 'WF': 'Wallis and Futuna', 'EH': 'Western Sahara', 'YE': 'Yemen', 'YU': 'Yugoslavia (former)',
    'ZM': 'Zambia', 'ZR': 'Zaire (former)', 'ZW': 'Zimbabwe'}

def readFile():
    # read in customer information csv file
    cust_info = pd.read_csv(os.environ['CSV_FILE'])
    # get necessary customer info columns from csv file
    cust_info = cust_info[['Name', 'Paid at', 'Billing Name', 'Shipping Country', 'Lineitem name', 'Total',
                         'Refunded Amount']]
    cust_info['Paid at'] = cust_info['Paid at'].str.slice(0, 10)
    cust_info['Paid at'] = pd.to_datetime(cust_info['Paid at'], format='%Y-%m-%d')

    # get handle on existing file and output customer info onto Excel sheet
    wb = load_workbook(filename=os.environ['EXCEL_SHEET'])
    # get current month and year worksheet
    month_year = datetime.date.today().strftime("%b-%Y").split("-")
    worksheet = "{month}. {year}".format(month=month_year[0], year=month_year[1])
    ws = wb[worksheet]
    font = Font(name='Times New Roman', size=12)
    alignment = Alignment(horizontal='center')
    # format to display currency symbol in front of price and '-' for $0 values
    # Note: This will not display the currency in Accounting format
    fmt_acct = u'$#,##0.00;;$  -;'

    sales_order_col = 'A'
    date_col = 'B'
    customer_col = 'C'
    country_col = 'D'
    product_col = 'E'
    total_sales_col = 'F'
    refunds_col = 'H'
    first_order_no = ws[sales_order_col + '2'].value[1:]
    print(first_order_no)

    for index, row in cust_info.iterrows():
        try:
            order_no = row['Name'][1:]
            excel_row = int(order_no) - int(first_order_no) + 2
            # print(order_no)
            # print(excel_row)
            # print(order_no, row['Paid at'].strftime("%m/%d/%Y"))
            # check to see there is no information about sales order
            if ws[sales_order_col + str(excel_row)].value != row['Name']:
                ws[sales_order_col + str(excel_row)] = row['Name']
                ws[sales_order_col + str(excel_row)].font = font
                ws[sales_order_col + str(excel_row)].alignment = alignment
                ws[date_col + str(excel_row)] = row['Paid at'].strftime("%m/%d/%Y")
                ws[date_col + str(excel_row)].font = font
                ws[date_col + str(excel_row)].alignment = alignment
                cust_name = row['Billing Name'].strip().lower().title().split()
                ws[customer_col + str(excel_row)] = " ".join(cust_name)
                ws[customer_col + str(excel_row)].font = font
                ws[customer_col + str(excel_row)].alignment = alignment
                ws[country_col + str(excel_row)] = country_abbreviations[row['Shipping Country']]
                ws[country_col + str(excel_row)].font = font
                ws[country_col + str(excel_row)].alignment = alignment
                ws[product_col + str(excel_row)] = row['Lineitem name']
                ws[product_col + str(excel_row)].font = font
                ws[product_col + str(excel_row)].alignment = alignment
                ws[total_sales_col + str(excel_row)] = float(row['Total'])
                ws[total_sales_col + str(excel_row)].font = font
                ws[total_sales_col + str(excel_row)].number_format = fmt_acct
                ws[total_sales_col + str(excel_row)].alignment = alignment
            # double check information and see if order has been refunded
            ws[refunds_col + str(excel_row)] = float(row['Refunded Amount'])
            ws[refunds_col + str(excel_row)].font = font
            ws[refunds_col + str(excel_row)].number_format = fmt_acct
            ws[refunds_col + str(excel_row)].alignment = alignment
        # continue onto the next row if some columns are empty
        except ValueError:
            continue

    wb.save(filename=os.environ['EXCEL_SHEET'])

if __name__ == '__main__':
    readFile()

